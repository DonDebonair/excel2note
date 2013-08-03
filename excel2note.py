#!/usr/bin/env python

import time
import argparse
from xml.sax.saxutils import escape
from openpyxl import load_workbook


def excel_2_enex(excel_file, enex_file, name_row=0, title_column=0):

    of = open(enex_file, "wb")
    of.write('<?xml version="1.0" encoding="UTF-8"?>\n')
    of.write('<!DOCTYPE en-export SYSTEM "http://xml.evernote.com/pub/evernote-export.dtd">\n')
    of.write('<en-export export-date="' + time.strftime('%Y%m%dT%H%M%SZ') + '" application="excel2enex" version="0.5">\n')

    wb = load_workbook(filename=excel_file)
    sheet = wb.worksheets[0]

    column_names = []
    i = 0
    while True:
        cell_contents = sheet.cell(row=name_row, column=i).value
        if cell_contents is None:
            break
        column_names.append(cell_contents)
        i += 1

    max_row = sheet.get_highest_row()
    for row in xrange(name_row + 1, max_row):
        of.write('<note>')
        title = sheet.cell(row=row, column=title_column).value
        of.write('<title>' + escape(str(title)) + '</title>')
        of.write('<content>')
        of.write('<![CDATA[<?xml version="1.0" encoding="UTF-8"?><!DOCTYPE en-note SYSTEM "http://xml.evernote.com/pub/enml2.dtd">')
        of.write('<en-note>')
        for column in xrange(0, len(column_names) - 1):
            if column != title_column:
                cell_contents = sheet.cell(row=row, column=column).value
                of.write('<strong>%(column)s:</strong> %(value)s<br />' % {'column': column_names[column], 'value': escape(str(cell_contents))})
        of.write('</en-note>]]>')
        of.write('</content>')
        ctime = time.strftime('%Y%m%dT%H%M%SZ')
        mtime = time.strftime('%Y%m%dT%H%M%SZ')
        of.write('<created>' + ctime + '</created>')
        of.write('<updated>' + mtime + '</updated>')
        of.write('</note>\n')

    of.write('</en-export>\n')
    of.close()

    return max_row - name_row - 1


def main():
    parser = argparse.ArgumentParser(description="Convert an Excel sheet to a collection of Evernote notes contained in an .enex file", prog="excel2note")
    parser.add_argument("excel_file", help="The Excel file you want to convert")
    parser.add_argument("enex_file", help="The Evernote file you want to create")
    parser.add_argument("-n", "--names", help="The row containing the column names (defaults to 1)", type=int, default=1, metavar="<row nr>")
    parser.add_argument("-t", "--title", help="The column containing the text to be used as note-titles (defaults to 1)", type=int, default=1, metavar="<column nr>")
    args = parser.parse_args()
    rows = excel_2_enex(args.excel_file, args.enex_file, args.names - 1, args.title - 1)
    print "Processed %(rows)i rows" % {'rows': rows}

if __name__ == '__main__':
    main()


